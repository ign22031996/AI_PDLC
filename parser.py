import openpyxl
import yaml
import io


def parse_xlsx(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== Лист: {sheet_name} ===\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)


def parse_markdown(content: bytes, filename: str) -> str:
    text = content.decode("utf-8", errors="replace")
    return f"\n=== Модуль: {filename} ===\n{text}"


def parse_checklist_yaml(content: bytes, filename: str) -> str:
    text = content.decode("utf-8", errors="replace")
    try:
        data = yaml.safe_load(text)
    except yaml.YAMLError:
        return f"\n=== Чек-лист: {filename} ===\n{text}"

    lines = [f"\n=== Чек-лист: {filename} ==="]

    name = data.get("name") or filename
    lines.append(f"Проект: {name}")

    for section in data.get("sections", []):
        section_name = section.get("name", "")
        if section_name:
            lines.append(f"\nРаздел: {section_name}")
        for q in section.get("questions", []):
            q_name = q.get("name", "")
            q_desc = q.get("description", "")
            q_rating = q.get("rating", "")
            lines.append(f"\n  Задание: {q_name} (баллы: {q_rating})")
            if q_desc:
                lines.append(f"  Критерии:\n{q_desc}")

    return "\n".join(lines)


def build_program_text(
    xlsx_bytes: bytes,
    md_files: list[tuple[str, bytes]],
    yaml_files: list[tuple[str, bytes]] | None = None,
) -> str:
    parts = []

    parts.append("=== ПАСПОРТ ОБРАЗОВАТЕЛЬНОЙ ПРОГРАММЫ ===")
    parts.append(parse_xlsx(xlsx_bytes))

    if md_files:
        parts.append("\n=== ОПИСАНИЕ МОДУЛЕЙ ПРОГРАММЫ ===")
        for filename, content in md_files:
            parts.append(parse_markdown(content, filename))

    if yaml_files:
        parts.append("\n=== ЧЕК-ЛИСТЫ ПРОЕКТОВ (критерии оценки заданий) ===")
        for filename, content in yaml_files:
            parts.append(parse_checklist_yaml(content, filename))

    return "\n".join(parts)
